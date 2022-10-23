import openpyxl
import unidecode

################################# DATASET & GLOBAL VARs DEFINITION #################################

wb = openpyxl.load_workbook("WCNeutras_respostas_full.xlsx")
ws = wb['Sheet1']

# B = ws['B']
# b_header = B[0].value

# instituto = {}

# for cell in B:
#     if cell.value == b_header or cell.value == 'Outra':
#         continue

#     if cell.value in instituto.keys():
#         continue

#     instituto[cell.value] = True

# # for k in instituto.keys():
# #     print(k)

# print("===")

C = ws['C']
c_header = C[0].value

other_instituto = {}


################################# FUNCTION DEFINITION #################################

# FUNCTION 
def filter_dataset(data = None, output_file = "institutos_res.txt"):
    
    others_list = list(other_instituto.keys())

    if data != None:
        for value in data:
            s = value_filtering(value)

            if s not in other_instituto.keys():
                other_instituto[s] = value
            
    if output_file != None:

        with open(output_file,"w") as f:
            f.write(str(others_list))

    return others_list
        
# FUNCTION 
def string_formatting(value):

    s = str(value).lower().strip()
    s = unidecode.unidecode(s)
    s = s.replace(":", "")
    s = s.replace(" da ", " ")
    s = s.replace(" do ", " ")
    s = s.replace(" de ", " ")
    s = s.replace(" - ", " ")
    s = s.replace("-", " ")
    s = s.replace(".", " ")
    s = s.replace("  ", " ")
    s = s.replace(",", "")

    return s

# FUNCTION 
def value_filtering(value):

    s = string_formatting(value)

    if s in mapping.keys():
        s = mapping[s]

    else:
        c = s.split(" ")

        for idx in range(len(c)):
            if c[idx] in mapping.keys():
                c[idx] = mapping[c[idx]]

        s = " ".join(c)

    return s

def map():

    mapping = {}
    mapping['up'] = string_formatting('universidade porto')
    #mapping['ul'] = string_formatting('universidade lisboa')
    mapping['iade'] = string_formatting('faculdade design tecnologia comunicacao')
    mapping['ismai'] = string_formatting('universidade maia')
    mapping['ips'] = string_formatting('instituto politecnico setubal')
    mapping['ufp'] = string_formatting('universidade fernando pessoa')
    mapping['isel'] = string_formatting('instituto superior engenharia lisboa')
    mapping['ispa'] = string_formatting('instituto superior psicologia aplicada')
    #mapping['faul'] = 'faculdade artes universidade lisboa'
    mapping['isep'] = string_formatting('instituto superior educacao personalizada')
    mapping['fcup'] = string_formatting('faculdade de ciencias universidade porto')
    mapping['university of lisbon'] = string_formatting('universidade lisboa')
    mapping['imm'] = string_formatting('instituto medicina molecular')
    mapping['escs'] = string_formatting('escola superior ciencias saude')
    mapping['unl'] = string_formatting('universidade nova lisboa')
    mapping['iscte'] = string_formatting('instituto universitario lisboa')
    mapping['nova'] = string_formatting('universidade nova lisboa')
    mapping['fpul'] = string_formatting('faculdade psicologia universidade lisboa')
    mapping['ulisboa'] = string_formatting('universidade lisboa')
    mapping['university'] = string_formatting('universidade')
    mapping['lisbon'] = string_formatting('lisboa')
    mapping['ubi'] = string_formatting('universidade beira interior')
    mapping['iul'] = string_formatting('instituto universitario de lisboa')
    mapping['ipc'] = string_formatting('instituto politecnico de coimbra')
    # !!!! key repetida
    mapping['faul'] = string_formatting('faculdade arquitectura universidade lisboa')
    mapping['sst'] = string_formatting('school science technology')

    return mapping


################################# MAIN #################################

mapping = map()

#count = 0
for cell in C:
    if cell.value == c_header or cell.value is None:
        continue

    s = value_filtering(cell.value)

    if s in other_instituto.keys():
        continue

    other_instituto[s] = cell.value
    #count+=1

# print ("{:<100} {:<100}".format('Original', 'Final'))
# for k,v in other_instituto.items():
#     print ("{:<100} {:<100}   ".format(v, k))

# print("===")
# print(count)

# ## TEST 
# print()
# print(filter_dataset())